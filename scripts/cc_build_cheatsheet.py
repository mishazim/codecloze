"""Builds CodeCloze answer cheat-sheet .xlsx (8 tabs: category x difficulty).
Reads the extractor JSON on stdin. Annotates each level with Big-O, with the
BEST-CASE time complexity called out first (the training target).
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# id -> (best-case time, worst-case time, space).  Standard textbook complexities;
# best-case is the training-focus column.
BIGO = {
    # ---- algorithms / easy ----
    "linear-search": ("O(1)", "O(n)", "O(1)"),
    "binary-search": ("O(1)", "O(log n)", "O(1)"),
    "bubble-sort": ("O(n)", "O(n^2)", "O(1)"),
    "insertion-sort": ("O(n)", "O(n^2)", "O(1)"),
    "selection-sort": ("O(n^2)", "O(n^2)", "O(1)"),
    "fibonacci-memoization": ("O(n)", "O(n)", "O(n)"),
    "dfs": ("O(V+E)", "O(V+E)", "O(V)"),
    "bfs": ("O(V+E)", "O(V+E)", "O(V)"),
    "inorder-traversal": ("O(n)", "O(n)", "O(h)"),
    "preorder-traversal": ("O(n)", "O(n)", "O(h)"),
    "postorder-traversal": ("O(n)", "O(n)", "O(h)"),
    "level-order-traversal": ("O(n)", "O(n)", "O(n)"),
    "two-pointers": ("O(1)", "O(n)", "O(1)"),
    "sliding-window-fixed": ("O(n)", "O(n)", "O(1)"),
    "xor-single-number": ("O(n)", "O(n)", "O(1)"),
    # ---- algorithms / medium ----
    "merge-sort": ("O(n log n)", "O(n log n)", "O(n)"),
    "quick-sort": ("O(n log n)", "O(n^2)", "O(log n)"),
    "counting-sort": ("O(n+k)", "O(n+k)", "O(k)"),
    "radix-sort": ("O(d(n+k))", "O(d(n+k))", "O(n+k)"),
    "sliding-window-variable": ("O(n)", "O(n)", "O(k)"),
    "monotonic-stack": ("O(n)", "O(n)", "O(n)"),
    "kadane": ("O(n)", "O(n)", "O(1)"),
    "floyd-cycle-detection": ("O(n)", "O(n)", "O(1)"),
    "kmp": ("O(n+m)", "O(n+m)", "O(m)"),
    "trie": ("O(m)", "O(m)", "O(ALPHABET*N)"),
    "lru-cache": ("O(1)", "O(1)", "O(capacity)"),
    "union-find": ("O(alpha(n)) ~ O(1)", "O(alpha(n))", "O(n)"),
    "kahn-topological-sort": ("O(V+E)", "O(V+E)", "O(V)"),
    "dijkstra": ("O((V+E) log V)", "O((V+E) log V)", "O(V)"),
    "boyer-moore-voting": ("O(n)", "O(n)", "O(1)"),
    "fisher-yates-shuffle": ("O(n)", "O(n)", "O(1)"),
    "topological-sort-dfs": ("O(V+E)", "O(V+E)", "O(V)"),
    "jump-game": ("O(n)", "O(n)", "O(1)"),
    # ---- algorithms / hard ----
    "lcs": ("O(n*m)", "O(n*m)", "O(n*m)"),
    "lis": ("O(n log n)", "O(n^2)", "O(n)"),
    "knapsack-01": ("O(n*W)", "O(n*W)", "O(n*W)"),
    "coin-change": ("O(n*amount)", "O(n*amount)", "O(amount)"),
    "edit-distance": ("O(n*m)", "O(n*m)", "O(n*m)"),
    "segment-tree": ("O(log n)", "O(log n)", "O(n)"),
    "fenwick-tree": ("O(log n)", "O(log n)", "O(n)"),
    "kruskal": ("O(E log E)", "O(E log E)", "O(V)"),
    "prim": ("O((V+E) log V)", "O((V+E) log V)", "O(V)"),
    "bellman-ford": ("O(V*E)", "O(V*E)", "O(V)"),
    "quick-select": ("O(n)", "O(n^2)", "O(1)"),
    "tarjan-scc": ("O(V+E)", "O(V+E)", "O(V)"),
    "a-star": ("O(E)", "O(b^d)", "O(V)"),
    "manacher": ("O(n)", "O(n)", "O(n)"),
    "rabin-karp": ("O(n+m)", "O(n*m)", "O(1)"),
    # ---- algorithms / expert ----
    "floyd-warshall": ("O(V^3)", "O(V^3)", "O(V^2)"),
    "ford-fulkerson": ("O(E * maxflow)", "O(E * maxflow)", "O(V^2)"),
    "n-queens": ("O(n!)", "O(n!)", "O(n)"),
    "hamiltonian-cycle": ("O(n!)", "O(n!)", "O(n)"),
    "permutation-generation": ("O(n * n!)", "O(n * n!)", "O(n)"),
    "subset-generation": ("O(n * 2^n)", "O(n * 2^n)", "O(n)"),
    "bloom-filter": ("O(k)", "O(k)", "O(m bits)"),
    "aho-corasick": ("O(n+m+z)", "O(n+m+z)", "O(m * ALPHABET)"),
    "suffix-array": ("O(n log n)", "O(n log n)", "O(n)"),
    "strassen": ("O(n^2.807)", "O(n^2.807)", "O(n^2)"),
    "matrix-chain-multiplication": ("O(n^3)", "O(n^3)", "O(n^2)"),
    # ---- data structures / easy (representative op) ----
    "dynamic-array": ("O(1) amortized append", "O(n) on resize", "O(n)"),
    "stack": ("O(1)", "O(1)", "O(n)"),
    "queue": ("O(1)", "O(1)", "O(n)"),
    "singly-linked-list": ("O(1) head insert", "O(n) search", "O(n)"),
    "doubly-linked-list": ("O(1) with node ref", "O(n) search", "O(n)"),
    "hash-map": ("O(1)", "O(n)", "O(n)"),
    "hash-set": ("O(1)", "O(n)", "O(n)"),
    "binary-tree": ("O(1) root", "O(n)", "O(n)"),
    "bst": ("O(log n) balanced", "O(n) skewed", "O(n)"),
    "string-builder": ("O(1) amortized append", "O(n) on resize", "O(n)"),
    # ---- data structures / medium ----
    "min-heap": ("O(1) peek", "O(log n) push/pop", "O(n)"),
    "deque": ("O(1)", "O(1)", "O(n)"),
    "circular-buffer": ("O(1)", "O(1)", "O(n)"),
    "graph-adjacency-list": ("O(1) add edge", "O(V+E) traverse", "O(V+E)"),
    "graph-adjacency-matrix": ("O(1) edge lookup", "O(V^2) traverse", "O(V^2)"),
    "linked-hash-map": ("O(1)", "O(n)", "O(n)"),
    "circular-linked-list": ("O(1) insert", "O(n) search", "O(n)"),
    "skip-list": ("O(log n)", "O(n)", "O(n)"),
    # ---- data structures / hard ----
    "avl-tree": ("O(log n)", "O(log n)", "O(n)"),
    "segment-tree-ds": ("O(log n)", "O(log n)", "O(n)"),
    "fenwick-tree-ds": ("O(log n)", "O(log n)", "O(n)"),
    "lfu-cache": ("O(1)", "O(1)", "O(capacity)"),
    "b-tree": ("O(log n)", "O(log n)", "O(n)"),
    "suffix-tree": ("O(m) search", "O(n) build", "O(n)"),
    "trie-ds": ("O(m)", "O(m)", "O(ALPHABET*N)"),
    # ---- data structures / expert ----
    "red-black-tree": ("O(log n)", "O(log n)", "O(n)"),
    "lsm-tree": ("O(1) write", "O(log n) read", "O(n)"),
    "bloom-filter-ds": ("O(k)", "O(k)", "O(m bits)"),
    "bitset": ("O(1) per bit", "O(n/w) range", "O(n/w)"),
    "union-find-ds": ("O(alpha(n)) ~ O(1)", "O(alpha(n))", "O(n)"),
    "suffix-array-ds": ("O(m log n) search", "O(n log n) build", "O(n)"),
}

CAT_LABEL = {"algorithms": "Algorithms", "dataStructures": "Data Structures"}
DIFFS = ["easy", "medium", "hard", "expert"]
DIFF_LABEL = {"easy": "Easy", "medium": "Medium", "hard": "Hard", "expert": "Expert"}
# terminal-ish header fills per difficulty
DIFF_FILL = {"easy": "1F7A33", "medium": "B8860B", "hard": "B22222", "expert": "7A1FA2"}
LANGS = [("python", "Python"), ("java", "Java"), ("cpp", "C++")]

levels = json.load(sys.stdin)
# group preserving order
groups = {}
for lv in levels:
    groups.setdefault((lv["category"], lv["difficulty"]), []).append(lv)

wb = Workbook()
wb.remove(wb.active)

mono = Font(name="Consolas", size=9)
hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
top_wrap = Alignment(vertical="top", wrap_text=True)
top = Alignment(vertical="top")
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["Level", "Description", "Best-Case Time", "Worst-Case Time", "Space"]
for _, lname in LANGS:
    HEADERS += [f"{lname} Solution", f"{lname} Blanks"]

missing = []
for cat in ["algorithms", "dataStructures"]:
    for diff in DIFFS:
        rows = groups.get((cat, diff), [])
        ws = wb.create_sheet(f"{CAT_LABEL[cat]} - {DIFF_LABEL[diff]}")
        ws.sheet_properties.tabColor = DIFF_FILL[diff]
        # header
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(1, c, h)
            cell.font = hdr_font
            cell.fill = PatternFill("solid", fgColor=DIFF_FILL[diff])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.freeze_panes = "A2"
        r = 2
        for lv in rows:
            best, worst, space = BIGO.get(lv["id"], ("?", "?", "?"))
            if lv["id"] not in BIGO:
                missing.append(lv["id"])
            ws.cell(r, 1, lv["name"]).font = Font(bold=True)
            ws.cell(r, 2, lv["description"])
            ws.cell(r, 3, best).font = Font(bold=True, color="1F7A33")
            ws.cell(r, 4, worst)
            ws.cell(r, 5, space)
            c = 6
            for key, _ in LANGS:
                data = lv["langs"].get(key, {})
                code = data.get("code", "(none)")
                blanks = data.get("blanks", [])
                blank_str = "\n".join(f"{i}: {a}" for i, a in enumerate(blanks))
                cc = ws.cell(r, c, code); cc.font = mono
                cb = ws.cell(r, c + 1, blank_str); cb.font = mono
                c += 2
            for cc in range(1, len(HEADERS) + 1):
                cell = ws.cell(r, cc)
                cell.alignment = top_wrap if cc in (2,) or cc >= 6 else top
                cell.border = border
            r += 1
        # widths
        widths = [22, 40, 18, 18, 14] + [60, 26] * 3
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

OUT = "reference/CodeCloze_Cheatsheet.xlsx"
wb.save(OUT)
print(f"Saved {OUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"Levels written: {sum(len(v) for v in groups.values())}")
print(f"Missing Big-O entries: {missing or 'none'}")
